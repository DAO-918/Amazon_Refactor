import re
from Tool.Tools_Init import startInit
import pandas as pd
from openpyxl import load_workbook

import urllib.parse

## ! 步骤一格式化链接

config = startInit()

def format_url(url:str):
    if 'aax-us-iad.' in url:
        url = url.replace('aax-us-iad', 'www')
    if 'sspa' in url and 'dp%2F' in url:
        asin_pattern = re.compile(r'dp%2F([A-Za-z0-9]{10})')
    elif 'www.amazon.com' in url or 'www.amazon.co.uk' in url:
        asin_pattern = re.compile(r'/dp/([A-Z0-9]{10})')
    else:
        return None, None, None
    match = asin_pattern.search(url)
    asin = match[1] if match else None
    parsed_url = urllib.parse.urlparse(url)
    domain = parsed_url.netloc
    domain_parts = domain.split('.')
    if len(domain_parts) >= 4:
        domain_suffix = f'{domain_parts[-2]}.{domain_parts[-1]}'
    else:
        domain_suffix = domain_parts[-1]
    domain_suffix_country_dict = {'com': 'us', 'co.uk': 'uk'}
    country = domain_suffix_country_dict.get(domain_suffix, domain_suffix)
    new_url = f'https://{domain}/dp/{asin}'
    return new_url, asin, country

def Initialize_Excel(sheet_ASIN_path):
    sheet_ASIN = pd.read_excel(sheet_ASIN_path, sheet_name='Sheet1')
    wb = load_workbook(sheet_ASIN_path, data_only=False)
    sheet_ASIN = wb['Sheet1']

    for row_index, row in enumerate( # type: ignore
        sheet_ASIN.iter_rows(min_row=2, values_only=True), start=2
    ):    
        link = str(row[0])
        asin = str(row[1])
        country = str(row[2])
        if link is None or link=='None':
            break
        
        Link,ASIN,Country = format_url(link)
        if Link is None:
            sheet_ASIN.cell(row=row_index, column=4).value = 'link error' # type: ignore
            continue
        sheet_ASIN.cell(row=row_index, column=1).value = Link # type: ignore
        sheet_ASIN.cell(row=row_index, column=2).value = ASIN # type: ignore
        sheet_ASIN.cell(row=row_index, column=3).value = Country # type: ignore

    wb.save(sheet_ASIN_path)

# 测试代码
if __name__ == '__main__':
    Initialize_Excel(r'D:\AutoRPA\产品信息\产品竞品\ASIN_Info-总asin.xlsx')