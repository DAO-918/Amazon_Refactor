from openpyxl import load_workbook

# 打开 Excel 文件，获取工作簿对象
wb = load_workbook('./Test/Test.xlsx')

# 根据 sheet 名字获取第一个工作表对象
ws = wb['Sheet1']

# 遍历第一列的单元格
for row in ws.iter_rows(min_row=1, min_col=1, max_col=1):

    cell = row[0]  # 获取当前行的第一个单元格对象
    
    # 判断当前单元格是否为合并单元格
    merged = False
    merged_range = None
    
    # 判断是否为合并单元格
    # cell.coordinate 则是该单元格的行列坐标，例如 'A1'，'B2' 等
    # ws.merged_cells 返回一个已合并单元格范围的列表，每个元素是一个 tuple，包含起始行号、终止行号、起始列号和终止列号。
    # 判断当前单元格是否在已合并单元格的范围内，如果在，则说明该单元格是一个合并的单元格。
    if cell.coordinate in ws.merged_cells:
        merge_type = 'This is a merged cell.'
    else:
        merge_type = 'This is not a merged cell.'
    
    # ws.merged_cells.ranges返回一个包含工作表中所有合并单元格范围的列表。
    for mr in ws.merged_cells.ranges:
        # 如果当前单元格坐标在合并单元格范围内，则为合并单元格
        if cell.coordinate in mr:
            merge_type = 'This is a merged cell.'
            merged = True
            merged_range = mr
            break
    if not merged:
        merge_type = 'This is not a merged cell.'

    # 输出单元格信息
    if merged:
        cell_content = ws[merged_range.start_cell.coordinate].value
        print(f'Cell "{cell.coordinate}" value={cell_content}, info: Merged cell size={merged_range.size}, position={cell.row, cell.column}, {merge_type}')
    else:
        print(f'Cell "{cell.coordinate}" value={cell.value}, info: Position={cell.row, cell.column}, {merge_type}')

    # 输出单元格的大小、位置和内容
    # Python的openpyxl库中的Cell对象确实没有'size'属性。'size'在这种上下文中没有意义，因为每个单元格的尺寸始终是一致的。
    # print(f'Cell "{cell.coordinate}" info: size={cell.size}, position={cell.row, cell.column}, {merge_type}')
