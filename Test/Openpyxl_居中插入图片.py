import decimal
from tempfile import NamedTemporaryFile
from openpyxl.drawing.image import Image
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border
from openpyxl.styles import PatternFill
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.utils.cell import get_column_letter

DEFAULT_FONT = Font()
DEFAULT_FILL = PatternFill()
DEFAULT_ALIGNMENT = Alignment()
DEFAULT_BORDER = Border()
DEFAULT_NUMBER_FORMATE = ''


class AddImage(Image):
    def __init__(self, img, max_width=None, max_height=None):
        super().__init__(img)

        self.max_width = max_width or self.height
        self.max_height = max_height or self.width
        self.r_width, self.r_height = self.height, self.width

    def image_resizer(self):
        aspect_ratio = self.r_width / self.r_height
        if aspect_ratio <= 2:
            target_size_width = self.max_height * aspect_ratio
            self.width, self.height = target_size_width, self.max_height
        else:
            target_size_height = self.max_width / aspect_ratio  # w:h     150:shr_image_height
            self.width, self.height = self.max_width, target_size_height
        return self


class EWorkbook(Workbook):

    def __init__(self, write_only=False, iso_dates=False):
        super().__init__(write_only=write_only, iso_dates=iso_dates)

    def create_sheet(self, title=None, index=None):
        new_ws = EWorksheet(parent=self, title=title)
        self._add_sheet(sheet=new_ws, index=index)
        return new_ws

    def temporary_save(self):
        with NamedTemporaryFile() as tmp:
            file_path = tmp.name + '.xlsx'
            self.save(file_path)
        return file_path


class EWorksheet(Worksheet):
    def __init__(self, parent, title=None):
        super().__init__(parent, title=title)

    def add_cell(self, row: int, col: int, value: [int, str, float] = None, **kwargs):
        """
        向单元格写入数据
        :param row: The beginning row（开始写入的行）
        :param col: The beginning col（开始写入的列）
        :param value: 向单元格写入的数据
        :return: None
        """
        if row < 1 or col < 1:
            raise ValueError("Row or column values must be at least 1")

        cell = self._get_cell(row, col)
        cell.font = kwargs.get('font', DEFAULT_FONT)
        cell.alignment = kwargs.get('alignment', DEFAULT_ALIGNMENT)
        cell.fill = kwargs.get('fill', DEFAULT_FILL)
        cell.border = kwargs.get('border', DEFAULT_BORDER)
        cell.number_format = kwargs.get('number_format', DEFAULT_NUMBER_FORMATE)

        if value is 0:
            cell.value = 0
        elif value is None:
            cell.value = None
        elif isinstance(value, decimal.Decimal):
            cell.value = decimal.Decimal(str(value))
        else:
            cell.value = str(value).strip()
        return cell

    def add_row(self, row: int, col: int, values: list = None, **kwargs):
        """
        向sheet写入一列数据
        :param row: The beginning row（开始写入的行）
        :param col: The beginning col（开始写入的列）
        :param values: 向单元格写入一列的数据
        :return: None
        """
        if not values and not isinstance(values, list):
            values = []
        for idx, val in enumerate(values):
            self.add_cell(row, col + idx, value=val, **kwargs)

    def add_col(self, row: int, col: int, values: list = None, **kwargs):
        """
        向sheet写入一行数据
        :param row: The beginning row（开始写入的行）
        :param col: The beginning col（开始写入的列）
        :param values: 向sheet写入一行的数据
        :return: None
        """
        if not values and not isinstance(values, list):
            values = []
        for idx, val in enumerate(values):
            self.add_cell(row + idx, col, value=val, **kwargs)

    def add_merge_row(self, row: int, col: int, v_pcs: int = None, value: [int, str, float] = None, **kwargs):
        """
        向单元格写入数据，并向垂直方向合并单元格
        :param row: The beginning row（开始写入的行）
        :param col: The beginning col（开始写入的列）
        :param v_pcs: 垂直合并的行数
        :param value: 向单元格写入数据
        :return: None
        """
        if not v_pcs and not isinstance(v_pcs, int):
            h_pcs = 0
        self.add_cell(row, col, value=value, **kwargs)
        self.merge_cells(start_row=row,
                        start_column=col,
                        end_row=row + v_pcs,
                        end_column=col)

    def add_merge_rows(self, row: int, col: int, v_pcs: int = None, values: list = None, **kwargs):
        """
        向sheet写入一行数据，并向垂直方向合并单元格
        :param row: The beginning row（开始写入的行）
        :param col: The beginning col（开始写入的列）
        :param v_pcs: 垂直合并的行数
        :param values: 向sheet写入一列的数据
        :return: None
        """
        if not values and not isinstance(values, list):
            values = []

        for idx, val in enumerate(values):
            column_letter = get_column_letter(col + idx)
            if isinstance(val, tuple) and len(val) == 2:
                self.column_dimensions[column_letter].width = val[1]
                self.add_merge_row(row, col + idx, v_pcs=v_pcs, value=val[0], **kwargs)
            else:
                self.add_merge_row(row, col + idx, v_pcs=v_pcs, value=val, **kwargs)

    def add_merge_col(self, row: int, col: int, h_pcs: int = None, value: [int, str, float] = None, **kwargs):
        """
        向单元格写入数据，并向水平方向合并单元格
        :param row: The beginning row（开始写入的行）
        :param col: The beginning col（开始写入的列）
        :param h_pcs: 水平合并的列数
        :param value: 单元格写入的数据
        :return: None
        """
        if not h_pcs and not isinstance(h_pcs, int):
            h_pcs = 0
        self.add_cell(row, col, value=value, **kwargs)
        self.merge_cells(start_row=row,
                        start_column=col,
                        end_row=row,
                        end_column=col + h_pcs)

    def add_merge_cols(self, row: int, col: int, h_pcs: int = None, values: list = None, **kwargs):
        """
        向sheet写入一列数据，并向水平方向合并单元格
        :param row: The beginning row（开始写入的行）
        :param col: The beginning col（开始写入的列）
        :param h_pcs: 水平合并的列数
        :param values: 向sheet写入一行的值
        :return: None
        """
        if not values and not isinstance(values, list):
            values = []
        for idx, val in enumerate(values):
            # column_letter = get_column_letter(col + idx)
            # if isinstance(val, tuple) and len(val) == 2:
            #     self.column_dimensions[column_letter].width = val[1]
            #     self.add_merge_col(row + idx, col, v_pcs=v_pcs, value=val[0], **kwargs)
            # else:
            #     self.add_merge_col(row + idx, col, v_pcs=v_pcs, value=val, **kwargs)
            self.add_merge_col(row + idx, col, h_pcs=h_pcs, value=val, **kwargs)

    def block_height(self, row, v_pcs=None):
        """
        v 垂直方向； h 水平方向
        单元格em换算像素 高比率=4/3
        单元格em换算像素 宽比率=200/25
        :param row: The beginning row of Block
        :param v_pcs: The number of rows occupied by Block
        :return: 该块中每个单元格的高，返回的是一个列表
        """
        if not v_pcs and not isinstance(v_pcs, int):
            v_pcs = 1

        # 一个block包含多个单元格，获取每个单元格的高度，并以列表返回
        block_height_detail = []
        for i in range(v_pcs):
            height = self.row_dimensions[row + i].height
            if not height:
                # 如果没有设置高，默认值却为None？
                self.row_dimensions[row + i].height = 16
                block_height_detail.append(int(16/0.761886))
            else:
                block_height_detail.append(int(height/0.761886))
        return block_height_detail

    def block_width(self, col, h_pcs=None):
        """
        v 垂直方向； h 水平方向
        单元格em换算像素 高比率=4/3
        单元格em换算像素 宽比率=200/25
        :param col: The beginning col of Block
        :param h_pcs: The number of cols occupied by Block
        :return: 该块中每个单元格的宽，返回的是一个列表
        """
        if not h_pcs and not isinstance(h_pcs, int):
            h_pcs = 1

        # 一个block包含多个单元格，获取每个单元格的宽，并以列表返回
        block_width_detail = []
        for i in range(h_pcs):
            column_letter = get_column_letter(col+i)
            width = self.column_dimensions[column_letter].width
            # print(width, end='-')
            block_width_detail.append(width * (72/9))
        return block_width_detail

    @staticmethod
    def _get_location(image_size, block_size_lst):
        """
        Current position calculation
        :param image_size: Width / Height of Image
        :param block_size_lst: Width / height of each cell in Block
        :return: 定位的单元格(行/列)，定位的单元格(宽/高)，定位的单元格需要偏移的量
        """
        total_size = 0
        current_size = 0
        current = 0
        block_off_size = abs(sum(block_size_lst)-image_size) / 2
        for idx, size in enumerate(block_size_lst):
            total_size += size
            current_size = size
            current = idx
            if total_size > block_off_size:
                break
        current_off_size = int(current_size - total_size + block_off_size)
        return current, current_size, current_off_size

    def add_relative_image(self, row, col, image_url, v_pcs=None, h_pcs=None, max_width=None, max_height=None):
        """
        v 垂直方向； h 水平方向
        Add an image to the sheet.
        Optionally provide a cell for the top-left anchor
        :param row: The beginning row of Block
        :param col: The beginning col of Block
        :param image_url: Image Path
        :param v_pcs: The number of rows occupied by Block
        :param h_pcs: The number of cols occupied by Block
        :param max_width: The maximum width of the Image
        :param max_height: The height width of the Image
        :return: None
        """
        try:
            img = AddImage(image_url, max_width=max_width, max_height=max_height).image_resizer()
            p2e = pixels_to_EMU
            img_size = XDRPositiveSize2D(p2e(img.height), p2e(img.width))
            block_height_lst = self.block_height(row, v_pcs=v_pcs)
            block_width_lst = self.block_width(col, h_pcs=h_pcs)
            current_row, current_row_height, current_off_height = self._get_location(img.width, block_height_lst)
            current_col, current_col_width, current_off_width = self._get_location(img.height, block_width_lst)
            marker = AnchorMarker(col=col + current_col - 1, colOff=p2e(current_off_width),
                                row=row + current_row - 1, rowOff=p2e(current_off_height))
            img.anchor = OneCellAnchor(_from=marker, ext=img_size)
            self.add_image(img)
        except Exception as e:
            pass


if __name__ == '__main__':
    import tkinter as tk
    from tkinter import filedialog
    '''打开选择文件夹对话框'''
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename()  # 获得选择好的文件
    wb = EWorkbook()
    ws = wb.create_sheet("sss", 0)
    ws.add_relative_image(12, 1, filepath, 5, 4, max_width=180, max_height=80)
    ws.cell(1, 2, "jacky")
    ws.add_row(2, 2, ["jacky", '4444', None, False, True, 0, '00000'])
    ws.add_col(3, 2, ["ccc", 'dddd444'])
    wb.save('2224.xlsx')