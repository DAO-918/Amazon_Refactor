import os
from openpyxl import load_workbook

# 加载Excel文件
wb = load_workbook('C:\\Users\\White\\Desktop\\鼠标垫SKU及图片链接(1).xlsx')
sheet = wb.active

# 创建字典保存文件名对照表数据
filename_dict = {}

# 读取每一行的数据
for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
    old_name, new_name = row
    filename_dict[old_name] = new_name

# 定义遍历文件夹的路径
base_dir = "K:\\80x30-SKU-4K"

# 遍历文件夹
for filename in os.listdir(base_dir):
    # 如果文件在我们的命名对照表中
    for old_name in filename_dict.keys():
        if filename.split('.')[0].split('_')[0] in old_name :
            #new_name = filename.replace(old_name, filename_dict[old_name])
            old_path = os.path.join(base_dir, filename)
            new_path = os.path.join(base_dir, old_name)
            os.rename(old_path, new_path)
            #break  # 一旦找到匹配并完成重命名，就跳出循环

