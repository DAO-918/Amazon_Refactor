from openpyxl import load_workbook

# 打开 Excel 文件，获取工作簿对象
wb = load_workbook('./Test/Test.xlsx')
# 根据 sheet 名字获取第一个工作表对象
sheet = wb['Sheet1']
row_count = 0
while sheet.cell(row=row_count+1, column=1).value is not None:
    row_count += 1
column_count = 0
while sheet.cell(row=1, column=column_count+1).value is not None:
    column_count += 1       
print(row_count, column_count)

print(sheet.max_row, sheet.max_column)

