import os
import openpyxl
import numpy as np
from PIL import Image
from openpyxl.drawing.image import Image as Img
from openpyxl.utils import get_column_letter
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils.units import pixels_to_EMU

from openpyxl.drawing.xdr import XDRPositiveSize2D


# 加载 Excel 文件
workbook_A = openpyxl.load_workbook(filename='./Test/A.xlsx')  
workbook_B = openpyxl.load_workbook(filename='./Test/B.xlsx')  
output_folder = './Test/IMG'  
# 选择要处理的工作表
worksheet_A = workbook_A['Sheet1']  
worksheet_B = workbook_B['Sheet1']  

worksheet_B.column_dimensions['A'].width = 8

# 获取工作表中的所有图像
images = worksheet_A._images
# 遍历图像并打印位置信息
for index, image in enumerate(images):
    # 获取图像的左上角行号，即图片所在行
    row = image.anchor.to.row + 1
    column = image.anchor.to.col + 1
    if row == 1:
        with open(os.path.join(output_folder, 'text.png'), 'wb') as img_file:
            img_pil = Image.open(image.ref).convert("RGB")
            img_pil.save(os.path.join(output_folder, 'text.png'))
        continue
    column_letter = get_column_letter(column)
    #if f"{column_letter}{row}" in worksheet_A.merged_cells:
    #    continue
    # 获取对应的图片名称
    pic_name = worksheet_A.cell(row=row, column=4).value
    # 保存图片到本地并按D列的图片名命名
    img_path = os.path.join(output_folder, f'{pic_name}.png')
    with open(os.path.join(output_folder, f'{pic_name}.png'), 'wb') as img_file:
            img_pil = Image.open(image.ref).convert("RGB")
            img_pil.save(img_path)
    # 插入图片到表B的A列（行号对应表A同样位置）
    img = Img(img_path)
    img.width = 63 # col_ch * 8  col_ch = 8
    img.height = 61 # row_pt * (4 / 3) row_pt = 46
    # 获取图片原始尺寸
    original_width, original_height = img.width, img.height
    # 维持原比例，最宽63
    '''if original_width > original_height and original_width > 63:
        scale_ratio = original_width / 63
        img.width = 63
        img.height = int(original_height / scale_ratio)
    # 维持原比例，最高61
    elif original_height >= original_width and original_height > 61:
        scale_ratio = original_height / 61
        img.height = 61
    img.width = int(original_width / scale_ratio)'''
    # 修改行高
    worksheet_B.row_dimensions[row].height = 46
    worksheet_B.cell(row=row, column=1, value="")
    # 获取单元格的宽度和高度
    cell_width = worksheet_B.column_dimensions[column_letter].width
    cell_height = worksheet_B.row_dimensions[row].height
    # 计算图片的原始宽度和高度
    original_width, original_height = img.width, img.height
    # 计算需要移动的距离以将图片居中
    #diff_x = (cell_width - original_width * 0.13) / 2   # 我们使用 ̃0.13将像素转换为EMU，这是一个近似值
    #diff_y = (cell_height - original_height) / 2
    #marker = AnchorMarker(col=1, colOff=pixels_to_EMU(diff_x), row=row - 1, rowOff=pixels_to_EMU(diff_y))
    #p2e = pixels_to_EMU
    #img_size = XDRPositiveSize2D(p2e(img.height), p2e(img.width))
    #img.anchor = OneCellAnchor(_from=marker, ext=None)

    #worksheet_B.add_image(img)
    worksheet_B.add_image(img, f"A{row}")
    # 在B列写入图片的本地路径
    worksheet_B.cell(row=row, column=2, value=img_path)

# 保存修改后的表格B
workbook_B.save('./Test/B.xlsx')  # 换成实际想保存的文件名
