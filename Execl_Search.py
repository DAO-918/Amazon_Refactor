from cgi import print_arguments
import os
import re
import json
from time import sleep
from datetime import datetime
from PIL import Image
from openpyxl.drawing.image import Image as Img
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from docx import Document

from selenium import webdriver
import selenium
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver import ActionChains
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support import expected_conditions

from bs4 import BeautifulSoup

from urllib.parse import urlparse

from docx import Document

import pyperclip
from pynput.keyboard import Controller, Key

import base64
from io import BytesIO
from PIL import Image
from PIL import Image

from Tool.Tool_Web import *
from Tool.Tool_Data import *
from Tool.Tool_SQL import *

class ExcelSearch:
    def __init__(self):
        # 获取当前目录的路径
        self.projectroot = os.path.dirname(os.path.abspath(__file__))
        parent_directory = os.path.dirname(self.projectroot)
        # 生成当前时间的字符串，格式为 YYYYMMDD_HHMM
        current_time = datetime.now().strftime("%Y%m%d_%H%M")
        # 设置文档文件名为当前时间
        docfilename = f"Execl_Read_Output_{current_time}.docx"
        output_root = os.path.join(parent_directory, '# OUTPUT', os.path.basename(self.projectroot))
        self.docfilepath = os.path.join(output_root, docfilename)
        self.doc = Document()
        
        source_root = os.path.join(parent_directory, '# 报价表整合')
        self.offer_root = os.path.join(source_root, '报价表')
        self.image_root = os.path.join(source_root, '图片库')
        self.gmatch_root = os.path.join(source_root, '匹配库', 'Google_Lens')
        
        self.e整合_path = os.path.join(source_root, '# 报价表整合.xlsx')
        self.e整合_wb = load_workbook(filename=self.e整合_path, read_only=False)
        self.e整合_Sheet1 = self.e整合_wb['Sheet1']
        self.e整合_colstr_图片 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='图片')
        self.e整合_colstr_命名方式 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='命名方式')
        self.e整合_colstr_图片路径 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='图片路径')
        self.e整合_colstr_图片名称 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='图片名称')
        self.e整合_colstr_来源 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='来源')
        self.e整合_colstr_品牌 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='品牌')
        self.e整合_colnum_图片 = column_index_from_string(self.e整合_colstr_图片)
        self.e整合_colnum_命名方式 = column_index_from_string(self.e整合_colstr_命名方式)
        self.e整合_colnum_图片路径 = column_index_from_string(self.e整合_colstr_图片路径)
        self.e整合_colnum_图片名称 = column_index_from_string(self.e整合_colstr_图片名称)
        self.e整合_colnum_品牌 = column_index_from_string(self.e整合_colstr_品牌)
        
        self.e目标报价表_path = None
        self.e目标报价表_wb = None
        
        self.options = webdriver.ChromeOptions()
        self.options.binary_location = r'D:\Code\chrome-win\chrome.exe'
        self.options.debugger_address = '127.0.0.1:9222'
        self.options.browser_version = '114.0.5734.0'
        self.service = Service(executable_path=r'D:\Code\chromedriver_win32\114\chromedriver.exe')
        
        # 定义程序路径和参数
        self.program_path = "D:\\Code\\chrome-win\\chrome.exe"
        self.program_args = [
            "--remote-debugging-port=9222",
            "--user-data-dir=E:\\Code\\selenium\\AutomationProfile 114 Seller 9222",
        ]
        self.shortcut_path = "D:\\Code\\chrome - New Selenium 2.lnk"
        
        self.driver = webdriver.Chrome(service=self.service, options=self.options)
        self.wait = WebDriverWait(self.driver, 30)
        self.actions = ActionChains(self.driver)
        
        # 生成当前时间的字符串，格式为 YYYYMMDD_HHMM
        current_time = datetime.now().strftime("%Y%m%d_%H%M")
        # 设置文档文件名为当前时间
        self.docfilename = f"Garb_Search_Output_{current_time}.docx"
        self.projectroot = os.path.dirname(os.path.abspath(__file__))
        parent_directory = os.path.dirname(self.projectroot)
        self.output_root = os.path.join(parent_directory, '# OUTPUT', os.path.basename(self.projectroot))
        self.docfilepath = os.path.join(self.output_root, self.docfilename)
        self.doc = Document()
        
        # 实例化键盘对象
        self.keyboard = Controller()

        # !找到列名对应的列序号，返回字母
    def find_colname_letter(self, sheet, rowindex, colname, match_mode='精准匹配'):
        # next：这个函数会返回一个迭代器的下一个元素。
        # next 用于获取满足条件（该行的值等于colname）的第一个元素的列字母。如果没有元素满足条件，它将返回一个默认值，这里是None
        '''return next(
            (
                cell.column_letter
                for cell in sheet[rowindex]
                if  (match_mode == '精准匹配' and cell.value == colname) or \
                    (match_mode == '模糊匹配' and colname in cell.value)
            ),
            None,)'''
        colstr = None
        for cell in sheet[rowindex]:
            if (match_mode == '精准匹配' and cell.value == colname) or (match_mode == '模糊匹配' and colname in cell.value):
                colstr = cell.column_letter
                return colstr
        if colstr is None:
            return None

    def start_chrome_program(self):
        os.startfile(self.shortcut_path)
        sleep(10)
        return True

    def open_page(self, valurl):
        # 打开valurl网页
        for i in range(1,5):
            new_driver = None
            try:
                print((f'driver.get({valurl})'))
                self.driver.switch_to.window(self.driver.window_handles[-1])
                self.driver.execute_script("window.open()") 
                self.driver.switch_to.window(self.driver.window_handles[-1])
                self.driver.get(valurl)
                break
            except Exception as e: 
                print(e)
                print(f'Exception occurred while getting page {valurl}, retrying {i+1} time')
                # 关闭可能打开的Chrome程序
                os.system('taskkill /f /im chrome.exe')
                # 启动新的Chrome程序
                self.start_chrome_program()
                sleep(2) # 等待2秒以确保程序启动
                new_driver = webdriver.Chrome(service=self.service, options=self.options)
                new_driver.maximize_window()
                # 如果driver启动失败，则抛出异常
                if new_driver is not None:
                    self.driver = new_driver
                else:
                    print(f"Failed to start driver within timeout: {i}.")
        # 1. 初始化WebDriverWait,设置最长等待时间为5秒:
        self.wait = WebDriverWait(self.driver, 30)
        # 2. 使用until方法设置等待条件:
        self.wait.until(EC.presence_of_all_elements_located((By.TAG_NAME, 'body')))

    def google_lens(self, filepath, savepath):
        results = []
        self.open_page("https://www.google.com")
        #google_button = self.driver.find_element(By.XPATH, '//*[@id="lensSearchButton"]')
        google_button = self.driver.find_element(By.XPATH, '/html/body/div[1]/div[3]/form/div[1]/div[1]/div[1]/div/div[3]/div[3]')
        self.actions.move_to_element(google_button)
        self.actions.click(google_button)
        self.actions.perform()
        #self.wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="uploadText"]')))
        #google_upload = self.driver.find_element(By.XPATH, '//*[@id="uploadText"]')
        self.wait.until(EC.presence_of_all_elements_located((By.XPATH, '//*[@id="ow19"]/div[3]/c-wiz/div[2]/div/div[3]/div[2]/div/div[2]/span')))
        google_upload = self.driver.find_element(By.XPATH, '//*[@id="ow19"]/div[3]/c-wiz/div[2]/div/div[3]/div[2]/div/div[2]/span')
        self.actions.move_to_element(google_upload)
        self.actions.click(google_upload)
        self.actions.perform()
        sleep(5)
        # 复制文件路径到剪贴板
        pyperclip.copy(filepath)
        # 模拟键盘按下Ctrl键,value就相当于+号,表示后面还有按键
        self.keyboard.press(Key.ctrl.value)
        self.keyboard.press('v')
        # 此时文件名的输入框已经完成了粘贴操作，需要释放掉ctrl键和v键
        self.keyboard.release(Key.ctrl.value)
        self.keyboard.release('v')
        # 模拟按下Enter键后释放Enter键，就大功告成了！
        self.keyboard.press(Key.enter)
        self.keyboard.release(Key.enter)
        # 截屏并处理
        screenshot = self.driver.get_screenshot_as_png()
        screenshot = Image.open(BytesIO(screenshot))
        # 截图 //*[@id="yDmH0d"]/c-wiz/div/div[2]/div/c-wiz/div/div[2]
        lens_total = self.driver.find_element(By.XPATH, '//*[@id="yDmH0d"]/c-wiz/div/div[2]/div/c-wiz/div/div[2]')
        # 截取特定元素的部分
        location = lens_total.location
        size = lens_total.size
        left = location['x']
        top = location['y']
        right = location['x'] + size['width']
        bottom = location['y'] + size['height']
        screenshot = screenshot.crop((left, top, right, bottom))
        screenshot.save(os.path.join(savepath, 'lens_total.png'))
        # 展示列 `//*[@id="yDmH0d"]/c-wiz/div/div[2]/div/c-wiz/div/div[2]/c-wiz/div/div/div/div[2]/div[1]/div/div/div/div[1]/div/div`
        # 点击 //*[@id="yDmH0d"]/c-wiz/div/div[2]/div/c-wiz/div/div[2]/c-wiz/div/div/div/div[2]/div[1]/div/div/div/div[1]/div/div/div[1]/div[1]/div/div/div[2]/div
        # `//div[@class='WF9wo' and text()='查看完全匹配的结果']`
        exact_match = self.driver.find_element(By.XPATH, "//div[@class='WF9wo' and text()='查看完全匹配的结果']")
        self.actions.move_to_element(exact_match)
        self.actions.click(exact_match)
        self.actions.perform()
        sleep(1)
        self.wait.until(EC.presence_of_all_elements_located((By.XPATH, "//div[text()='查看完全匹配的结果']")))
        match_ul = self.driver.find_element(By.XPATH, "//ul[@aria-label='包含所有结果的列表']")
        match_lis = match_ul.find_elements(By.TAG_NAME, 'li')
        # TODO
        images = []
        for li in match_lis:
            link = li.find_element(By.TAG_NAME, 'a').get_attribute('href')
            info = li.find_element(By.XPATH, './a/div/div/div/div')
            infos = info.find_elements(By.XPATH, './div')
            domain = infos[0].find_element(By.XPATH, './div[1]/div[2]').text
            title = infos[0].find_element(By.XPATH, './div[2]').text
            image = infos[1].find_element(By.TAG_NAME, 'img').get_attribute('src')
            result = [link, domain, title, image]
            results.append(result)
            images.append(image)
            # TODO是否先存储到表格中，再下载图片
            
        self.driver.close()
        return results

    def aibaba_lens(self):
        pass

    def loop_img_search(self):
        报价表整合_Sheet1_maxrow = self.e整合_Sheet1.max_row
        报价表整合_Sheet1_maxcol = self.e整合_Sheet1.max_column
        colstr_编号A = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='编号A')
        A_index_range = self.e整合_Sheet1[f'{colstr_编号A}{2}:{colstr_编号A}{报价表整合_Sheet1_maxcol}']
        A_index_dict = {}
        # 计算编号A各个前缀的最大编号
        for row_index, cell in enumerate(A_index_range, start=2):
            A_key = cell.value
            if A_key not in A_index_dict:
                A_index_dict[A_key] = [row_index]
            else:
                A_index_dict[A_key].append(row_index)
                
        for A_key, index_list in A_index_dict.items():
            output_folder = os.path.join(self.gmatch_root, A_key)
            if not os.path.exists(output_folder):
                os.makedirs(output_folder)
            for row_index in index_list:
                filepath = self.e整合_Sheet1.cell(row=row_index, column=self.e整合_colnum_图片路径).value
                filename = self.e整合_Sheet1.cell(row=row_index, column=self.e整合_colnum_图片名称).value
                match_output = os.path.join(output_folder, filename)
                if not os.path.exists(match_output):
                    os.makedirs(match_output)
                result = self.google_lens(filepath, match_output)
                
        
        for i, row in enumerate(self.e整合_Sheet1.iter_rows(min_row=2), start=1):
            filepath = row[self.e整合_colnum_图片路径 -1].value
            filename = row[self.e整合_colnum_图片名称 -1].value
            
            result = self.google_lens(filepath)


# 测试代码
if __name__ == '__main__':
    ex = ExcelSearch()
    ex.google_lens('D:\Code\# 报价表整合\图片库\玩具-积木\贝乐迪\贝乐迪_爱情系列_爱心(760PCS)_18148.png')
