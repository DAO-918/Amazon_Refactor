from cgi import print_arguments
import os
import re
import json
from datetime import datetime
from PIL import Image
from openpyxl.drawing.image import Image as Img
from openpyxl.styles import Alignment
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter
from docx import Document

class ExcelConver:
    def __init__(self):
        #self.s_name = os.path.splitext(source_file)[0]
        #self.t_name = os.path.splitext(target_file)[0]
        #self.s_wb = load_workbook(filename=source_file,read_only=True)
        #self.t_wb = load_workbook(filename=target_file)
        
        # 获取当前目录的路径
        self.projectroot = os.path.dirname(os.path.abspath(__file__))
        parent_directory = os.path.dirname(self.projectroot)
        # 生成当前时间的字符串，格式为 YYYYMMDD_HHMM
        current_time = datetime.now().strftime("%Y%m%d_%H%M")
        # 设置文档文件名为当前时间
        docfilename = f"Execl_Read_Output_{current_time}.docx"
        output_root = os.path.join(parent_directory, '# OUTPUT', os.path.basename(self.projectroot))
        self.docfilepath = os.path.join(output_root, docfilename)
        self.doc = Document()
        
        source_root = os.path.join(parent_directory, '# 报价表整合')
        self.offer_root = os.path.join(source_root, '报价表')
        self.image_root = os.path.join(source_root, '图片库')
        
        self.e整合_path = os.path.join(source_root, '# 报价表整合.xlsx')
        self.e整合_wb = load_workbook(filename=self.e整合_path, read_only=False)
        self.e整合_Sheet1 = self.e整合_wb['Sheet1']
        self.e整合_colstr_图片 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='图片')
        self.e整合_colstr_命名方式 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='命名方式')
        self.e整合_colstr_图片路径 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='图片路径')
        self.e整合_colstr_图片名称 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='图片名称')
        self.e整合_colstr_来源 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='来源')
        self.e整合_colstr_品牌 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='品牌')
        self.e整合_colnum_图片 = column_index_from_string(self.e整合_colstr_图片)
        self.e整合_colnum_命名方式 = column_index_from_string(self.e整合_colstr_命名方式)
        self.e整合_colnum_图片路径 = column_index_from_string(self.e整合_colstr_图片路径)
        self.e整合_colnum_图片名称 = column_index_from_string(self.e整合_colstr_图片名称)
        self.e整合_colnum_品牌 = column_index_from_string(self.e整合_colstr_品牌)
        
        self.e记录_path = os.path.join(source_root, '报价表记录.xlsx')
        self.e记录_wb = load_workbook(filename=self.e记录_path, read_only=False)
        self.e记录_Sheet1 = self.e记录_wb['Sheet1']
        self.e记录_colstr_报价表名称 = self.find_colname_letter(sheet=self.e记录_Sheet1, rowindex=1, colname='报价表名称')
        self.e记录_colstr_品牌 = self.find_colname_letter(sheet=self.e记录_Sheet1, rowindex=1, colname='品牌')
        self.e记录_colstr_类别 = self.find_colname_letter(sheet=self.e记录_Sheet1, rowindex=1, colname='类别')
        self.e记录_colstr_列名行号1 = self.find_colname_letter(sheet=self.e记录_Sheet1, rowindex=1, colname='列名行号1')
        self.e记录_colstr_列名行号2 = self.find_colname_letter(sheet=self.e记录_Sheet1, rowindex=1, colname='列名行号2')
        self.e记录_colstr_起始位置 = self.find_colname_letter(sheet=self.e记录_Sheet1, rowindex=1, colname='起始位置')
        self.e记录_colstr_记录时间 = self.find_colname_letter(sheet=self.e记录_Sheet1, rowindex=1, colname='记录时间')
        
        self.e对照_path = os.path.join(source_root, '报价表对照.xlsx')
        self.e对照_wb = load_workbook(filename=self.e对照_path, read_only=True)
        self.e对照_Sheet1 = self.e对照_wb['Sheet1']
        self.e对照_colstr_报价表整合列名 = self.find_colname_letter(sheet=self.e对照_Sheet1, rowindex=1, colname='报价表整合列名')
        self.e对照_colstr_是否匹配 = self.find_colname_letter(sheet=self.e对照_Sheet1, rowindex=1, colname='是否匹配')
        self.e对照_colstr_A精准匹配 = self.find_colname_letter(sheet=self.e对照_Sheet1, rowindex=1, colname='A精准匹配')
        self.e对照_colstr_A模糊匹配 = self.find_colname_letter(sheet=self.e对照_Sheet1, rowindex=1, colname='A模糊匹配')
        self.e对照_colstr_B精准匹配 = self.find_colname_letter(sheet=self.e对照_Sheet1, rowindex=1, colname='B精准匹配')
        self.e对照_colstr_B模糊匹配 = self.find_colname_letter(sheet=self.e对照_Sheet1, rowindex=1, colname='B模糊匹配')
        
        self.e目标报价表_path = None
        self.e目标报价表_wb = None

    def format_string(self, s):
        if s is not None and isinstance(s, str):
            return s.replace("（","(").replace("）",")").replace("，",",").replace("[","").replace("]","").replace("'","").replace("\"","").replace(" ","").replace("\n","").replace("\r","").replace("\t","").replace(" ","").upper()
        else:
            return s

    # !计算表格的总行数和总列数
    # 不如直接使用工作表的行数和列数
    def tool_count(self,sheet):
        row_count = 0
        while sheet.cell(row=row_count+1, column=1).value is not None:
            row_count += 1
        column_count = 0
        while sheet.cell(row=1, column=column_count+1).value is not None:
            column_count += 1       
        return row_count, column_count

    # !目标报价表：找到图片的位置和导出图片
    def excel_img_read(self):
        # 选择要处理的工作表
        worksheet = self.s_wb['Sheet2']
        # 获取工作表中的所有图像
        images = worksheet._images
        image_positions = {}
        # 遍历图像并打印位置信息
        for index, image in enumerate(images):
            # 获取图像的左上角和右下角坐标
            rownum = image.anchor.to.row
            colnum = image.anchor.to.col
            pic_name = f"{self.s_name}_{index}_{rownum}_{colnum}.png"  # 假设图片格式为PNG
            image_positions[pic_name] = (rownum, colnum)
            # data = image.ref
            with open(os.path.join(self.image_root, pic_name), 'wb') as img_file:
                # 打开文件路径为image.ref的图像文件。convert("RGB")被用来将图像转换为RGB格式。
                img_pil = Image.open(image.ref).convert("RGB")
                # 在这个NumPy数组中，图像被存储为一个三维数组，
                # 第一个维度代表图像的高度，第二个维度代表宽度，
                # 第三个维度代表颜色通道（在RGB图像中，有红、绿、蓝三个通道）。每个元素代表在该坐标位置的像素值。
                # img_array = np.array(img_pil)
                # 将 NumPy 数组保存为图像文件
                img_pil.save(os.path.join(self.image_root, pic_name))
                # 记录图片位置信息（这里假设仍需要记录）
                image_positions[pic_name] = (rownum, colnum)
        print(image_positions)

    # !报价表对照：格式化报价表对照的JSON数组
    def stand_excel(self):
        print('开始处理：报价表对照')
        # 加载工作簿并将其设置为可写
        e对照_wb_write = load_workbook(filename=self.e对照_path, read_only=False)
        # 获取工作簿的第一个工作表
        e对照_sheet1 = e对照_wb_write['Sheet1']
        # 遍历工作表中的所有行，从第二行开始（因为第一行是标题），到第六列
        for row in e对照_sheet1.iter_rows(min_row=2, max_col=6):
            # 遍历每一行中的每一个单元格
            for i, cell in enumerate(row, start=1):
                # 如果是第一列和第二列，跳过
                if i<= 2:
                    continue
                # 获取单元格的值
                cell_value = cell.value
                # 如果单元格的值是空的，跳过
                if cell_value is None:
                    continue
                # 尝试将单元格的值加载为JSON
                try:
                    json.loads(cell_value)
                # 如果加载失败，将值转换为JSON字符串
                except Exception:
                    # 将值从字符串转换为列表
                    new_value = self.format_string(cell_value).split(",")
                    # 将列表转换为JSON字符串，并使用中文编码
                    cell.value = json.dumps(new_value, ensure_ascii=False)  # convert list to json string with Chinese characters
        # 保存工作簿
        e对照_wb_write.save(self.e对照_path)

    # !报价表对照：读取值
    def read_excel_contrast(self):
        ws = self.c_wb['Sheet1']
        # 使用 Worksheet.iter_rows() A1,B1,C1,A2,B2,C2
        for row in ws.iter_rows(min_row=2, max_col=6):
            报价表整合列名 = row[0].value
            是否匹配 = row[1].value
            A精准匹配 = json.loads(row[2].value)
            A模糊匹配 = json.loads(row[3].value)
            B精准匹配 = json.loads(row[4].value)
            B模糊匹配 = json.loads(row[5].value)
            print(报价表整合列名, 是否匹配, A精准匹配, A模糊匹配, B精准匹配, B模糊匹配)

    # !目标报价表：找到目标报价表的单元格的值
    # 1.遍历指定行（col_rowindex）的所有列，寻找名称和colname匹配的列。匹配的方式取决于在match_mode中指定的模式
    # 2.在指定的行中找不到任何匹配列，函数将返回一个错误信息。
    # 3.如果找到了匹配的列，则返回位于value_rowindex行、匹配列中的单元格的值。
    def find_colname_rowindex_value(self, sheet, rowindex, colname, value_rowindex, match_mode='精准匹配'):        
        # 遍历指定行（col_rowindex）的所有列，寻找名称和colname匹配的列
        col_index = None
        for i in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=rowindex, column=i).value)
            # 如果单元格值为空，则跳出循环
            if cell_value is None:
                break
            # 如果匹配模式为精准匹配，则比较单元格值是否与colname相等
            # 如果匹配模式为模糊匹配，则比较colname是否在单元格值中
            elif (match_mode == '精准匹配' and cell_value == colname) or \
                (match_mode == '模糊匹配' and colname in cell_value):
                # 如果找到匹配列，则跳出循环
                col_index = i
                break
        # 如果没有找到匹配列，则返回None
        if col_index is None:
            return None
        # 返回位于value_rowindex行、匹配列中的单元格的值
        return sheet.cell(row=value_rowindex, column=col_index).value

    # !找到列名对应的列序号，返回字母
    def find_colname_letter(self, sheet, rowindex, colname, match_mode='精准匹配'):
        # next：这个函数会返回一个迭代器的下一个元素。
        # next 用于获取满足条件（该行的值等于colname）的第一个元素的列字母。如果没有元素满足条件，它将返回一个默认值，这里是None
        '''return next(
            (
                cell.column_letter
                for cell in sheet[rowindex]
                if  (match_mode == '精准匹配' and cell.value == colname) or \
                    (match_mode == '模糊匹配' and colname in cell.value)
            ),
            None,)'''
        colstr = None
        for cell in sheet[rowindex]:
            if (match_mode == '精准匹配' and cell.value == colname) or (match_mode == '模糊匹配' and colname in cell.value):
                colstr = cell.column_letter
                return colstr
        if colstr is None:
            return None

    # !遍历目标文件夹下对应后缀的文件
    def list_files_by_type(self, directory , file_type='.xlsx'):
        excel_files = []
        # 对指定目录及其所有子目录进行遍历
        for root, dirs, files in os.walk(directory):
            for file in files:
                if file.endswith(file_type):
                    excel_files.append(os.path.join(root, file))
        return excel_files

    # !报价表记录：记录报价表文件夹下所有.xlsx文件的位置信息
    def recode_excel(self):
        print('开始处理：报价表记录')
        # 获取文件列表，通过类型过滤
        file_list = self.list_files_by_type(self.offer_root, '.xlsx')
        
        # 遍历文件列表
        for file_path in file_list:
            # 获取文件名
            ase_name = os.path.basename(file_path)
            # 获取文件夹名
            dir_name = os.path.basename(os.path.dirname(file_path))
            # 获取父文件夹名
            pdir_name = os.path.basename(os.path.dirname(os.path.dirname(file_path)))
            # 检查表格中是否有至少一个单元格的值等于ase_name
            ase_name_exist = any(
                cell.value == ase_name
                for cell in self.e记录_Sheet1[self.e记录_colstr_报价表名称]
            )
            # 如果表格中没有符合条件的单元格，则进行插入操作
            if not ase_name_exist:
                # 获取 Sheet1 中的最大行和列
                #报价表记录_Sheet1_maxrow, 报价表记录_Sheet1_maxcol = self.tool_count(self.e记录_Sheet1)
                报价表记录_Sheet1_maxrow = self.e记录_Sheet1.max_row
                # 插入数据
                self.e记录_Sheet1[f'{self.e记录_colstr_报价表名称}{报价表记录_Sheet1_maxrow+1}'] = ase_name
                self.e记录_Sheet1[f'{self.e记录_colstr_品牌}{报价表记录_Sheet1_maxrow+1}'] = dir_name
                self.e记录_Sheet1[f'{self.e记录_colstr_类别}{报价表记录_Sheet1_maxrow+1}'] = pdir_name
                
        # 保存修改后的表格
        self.e记录_wb.save(self.e记录_path)

    def contrast_data(self):
        print('开始处理：报价表整合')
        # *打开"报价表记录"并计算它的最大行和最大列
        #报价表记录_Sheet1_maxrow, 报价表记录_Sheet1_maxcol = self.tool_count(self.e记录_Sheet1)
        报价表记录_Sheet1_maxrow = self.e记录_Sheet1.max_row
        报价表记录_Sheet1_maxcol = self.e记录_Sheet1.max_column
        # *遍历该表的每一行，如果某行的"记录时间"为None，则取出当前行的"报价表名称"，"品牌"，"类别"等信息，以及另一张表中的"列名行号1"，"列名行号2"，"起始位置"和"记录时间"相关信息。
        for e记录行号 in range(2, 报价表记录_Sheet1_maxrow + 1):
            if self.e记录_Sheet1[f'{self.e记录_colstr_记录时间}{e记录行号}'].value is not None:
                continue
            报价表名称 = self.e记录_Sheet1[f'{self.e记录_colstr_报价表名称}{e记录行号}'].value
            品牌 = self.e记录_Sheet1[f'{self.e记录_colstr_品牌}{e记录行号}'].value
            类别 = self.e记录_Sheet1[f'{self.e记录_colstr_类别}{e记录行号}'].value
            # 如果列名有两行，列名行号1 列名行号2 都是有值的
            列名行号1 = self.e记录_Sheet1[f'{self.e记录_colstr_列名行号1}{e记录行号}'].value
            列名行号2 = self.e记录_Sheet1[f'{self.e记录_colstr_列名行号1}{e记录行号}'].value
            起始位置 = self.e记录_Sheet1[f'{self.e记录_colstr_起始位置}{e记录行号}'].value
            记录时间 = self.e记录_Sheet1[f'{self.e记录_colstr_记录时间}{e记录行号}'].value
            # *如果"列名行号1"，"列名行号2"，"起始位置"为None，或者"记录时间"不为None，则跳过当前行的后续操作。
            if 列名行号1 is None or 列名行号2 is None or 起始位置 is None or 记录时间 is not None:
                continue
            # *如果不满足条件，则打开报价表，计算"报价表整合"的最大行和最大列，然后用这些信息以及从"报价表记录"中获取到的名称和品牌等信息，生成报价表的路径，并打开报价表。
            报价表路径 = os.path.join(self.offer_root, 类别, 品牌, 报价表名称)
            wb = load_workbook(filename=报价表路径, read_only=False)
            
            # *遍历报价表的每一个sheet，获取sheet的最大行和最大列，并且找到"图片"，"系列"，"名称"，"货号"等字段所在的列的列号。
            for sheetname in wb.sheetnames:
                # *打开每个sheet，计算"报价表整合"的最大行和最大列
                #报价表整合_Sheet1_maxrow, 报价表整合_Sheet1_maxcol = self.tool_count(self.e整合_Sheet1)
                报价表整合_Sheet1_maxrow = self.e整合_Sheet1.max_row
                ws = wb[sheetname]
                #ws_maxrow, ws_maxcol = self.tool_count(ws)
                ws_maxrow = ws.max_row
                ws_maxcol = ws.max_column
                # 如果sheet是空表
                if ws_maxrow <= 列名行号1 and ws_maxrow<= 列名行号2:
                    continue
                图片列号 = None
                系列列号 = None
                名称列号 = None
                货号列号 = None
                col_maxlen = 0
                
                # *遍历每个sheet的每一列，逐行检查和获取"报价表整合"列名，"A精准匹配"，"A模糊匹配"，"B精准匹配"，"B模糊匹配"等信息，
                # *然后根据这些信息找到对应"报价表整合"的列名，并将这些信息抄写入"报价表整合"表中。
                for i, col in enumerate(ws.iter_cols(min_row=1, max_row=ws_maxrow ,max_col=ws_maxcol), start=1):
                    col_letter = get_column_letter(i)
                    # iter_cols获取的是数组，起始是0，而表格中展示的是从1开始
                    ws_列名1 = col[列名行号1 - 1]
                    ws_列名2 = col[列名行号2 - 1]                        
                    ws_列名_value = None
                    ws_列名1_value = None
                    ws_列名2_value = None
                    in_one_row = 列名行号1 == 列名行号2
                    # 如果只有一行
                    find_flag = False
                    A_find_flag = False
                    B_find_flag = False
                    报价表整合列名 = None
                    # 记录Sheet页中row的最大值
                    if col_maxlen < len(col):
                        col_maxlen = len(col)
                    # 先判断是否同一行，是否是合并列表，获取单元格的值
                    if in_one_row:
                        ws_列名_value = ws_列名1.value
                    else: 
                        ws_列名1_merged = ws_列名1.coordinate in ws.merged_cells
                        ws_列名1_merged_range =  None
                        if ws_列名1_merged:
                            for mr in ws.merged_cells.ranges:
                            # 如果当前单元格坐标在合并单元格范围内，则为合并单元格
                                if ws_列名1.coordinate in mr:
                                    ws_列名1_merged_range = mr
                                    break
                            ws_列名1_value = ws[ws_列名1_merged_range.start_cell.coordinate].value
                        else:
                            ws_列名1_value = ws_列名1.value
                        ws_列名2_merged = ws_列名2.coordinate in ws.merged_cells
                        ws_列名2_merged_range =  None
                        if ws_列名2_merged:
                            for mr in ws.merged_cells.ranges:
                            # 如果当前单元格坐标在合并单元格范围内，则为合并单元格
                                if ws_列名2.coordinate in mr:
                                    ws_列名2_merged_range = mr
                                    break
                            ws_列名2_value = ws[ws_列名2_merged_range.start_cell.coordinate].value
                        else:
                            ws_列名2_value = ws_列名2.value
                        # 判断是否上下相同
                        if ws_列名1_value == ws_列名2_value:
                            ws_列名_value = ws_列名1_value
                        # 格式化内容
                        if ws_列名_value is not None:
                            ws_列名_value = self.format_string(ws_列名_value)
                        elif ws_列名1_value is not None and ws_列名2_value is not None and ws_列名_value is None:
                            ws_列名1_value = self.format_string(ws_列名1_value)
                            ws_列名2_value = self.format_string(ws_列名2_value)
                    # 找到列对应报价表整合中的列名
                    # 即找到上面循环中的列名在报价表整合中对应的列名
                    for index, 报价表对照_row in enumerate(self.e对照_Sheet1.iter_rows(min_row=2, max_col=7)):
                        报价表整合列名 = 报价表对照_row[column_index_from_string(self.e对照_colstr_报价表整合列名)-1].value
                        是否匹配 = 报价表对照_row[column_index_from_string(self.e对照_colstr_是否匹配)-1].value
                        if 是否匹配 == '0':
                            continue
                        # ?如果单元格未空，json.loads返回
                        if 报价表对照_row[column_index_from_string(self.e对照_colstr_A精准匹配)-1].value is not None:
                            A精准匹配 = json.loads(报价表对照_row[column_index_from_string(self.e对照_colstr_A精准匹配)-1].value)
                        if 报价表对照_row[column_index_from_string(self.e对照_colstr_A模糊匹配)-1].value is not None:  
                            A模糊匹配 = json.loads(报价表对照_row[column_index_from_string(self.e对照_colstr_A模糊匹配)-1].value)
                        if 报价表对照_row[column_index_from_string(self.e对照_colstr_B精准匹配)-1].value is not None:
                            B精准匹配 = json.loads(报价表对照_row[column_index_from_string(self.e对照_colstr_B精准匹配)-1].value)
                        if 报价表对照_row[column_index_from_string(self.e对照_colstr_B模糊匹配)-1].value is not None:
                            B模糊匹配 = json.loads(报价表对照_row[column_index_from_string(self.e对照_colstr_B模糊匹配)-1].value)
                        # 如果是同一行 或者 合并单元格的两行值相同
                        if ws_列名_value is not None:
                            if find_flag == False and A精准匹配 is not None:
                                for exact in A精准匹配:
                                    if exact == ws_列名_value:
                                        find_flag = True
                                        break
                            if find_flag == False and A模糊匹配 is not None:
                                for exact in A模糊匹配:
                                    if exact in ws_列名_value:
                                        find_flag = True
                                        break
                        # 如果不是同一行 且 合并单元格的两行值不同
                        elif ws_列名1_value is not None and ws_列名2_value is not None and ws_列名_value is None:
                            if A_find_flag == False and A精准匹配 is not None:
                                for exact in A精准匹配:
                                    if exact == ws_列名1_value:
                                        A_find_flag = True
                                        break
                            if A_find_flag == False and A模糊匹配 is not None:
                                for exact in A模糊匹配:
                                    if exact in ws_列名1_value:
                                        A_find_flag = True
                                        break
                            if B_find_flag == False and B精准匹配 is not None:
                                for exact in B精准匹配:
                                    if exact == ws_列名2_value:
                                        B_find_flag = True
                                        break
                            if B_find_flag == False and B模糊匹配 is not None:
                                for exact in B模糊匹配:
                                    if exact in ws_列名2_value:
                                        B_find_flag = True
                                        break
                        if find_flag == True or (A_find_flag == True and B_find_flag == True):
                            break
                    # 上一次循环后如果匹配成功，则找到了对应的列名，记录后面需要的列号
                    if find_flag == True or (A_find_flag == True and B_find_flag == True):
                        if 报价表整合列名 == '图片':
                            图片列号 = i
                        if 报价表整合列名 == '系列':
                            系列列号 = i
                        if 报价表整合列名 == '名称':
                            名称列号 = i
                        if 报价表整合列名 == '货号': 
                            货号列号 = i
                    else:
                        # 如果没有匹配成功，跳出本次循环
                        continue
                    # 定位目标列，然后把值写到报价表整合中
                    目标列 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname=报价表整合列名)
                    目标列_colnum = column_index_from_string(目标列)
                    # col 是报价表循环的其中一列，是数组 0指第一行 起始位置是行号要减一
                    # 报价表整合_Sheet1_maxrow 不能在现在这个循环内得出
                    position = 1
                    # range() 函数生成的范围包含起始值，但不包含结束值
                    for row_index in range(int(起始位置-1), len(col)):
                        # 单元格为空值时, cell.value=None, 那么再写入其他表格中, 在表格中显示的是什么
                        print(f'写入数据: {报价表整合列名} 列\t {报价表整合_Sheet1_maxrow + position} 行')
                        写入数据 = self.format_string(col[row_index].value)
                        self.e整合_Sheet1.cell(row=报价表整合_Sheet1_maxrow + position, column=目标列_colnum, value=写入数据)
                        position = position +1
                    # openpyxl 中，不能直接为一个范围的单元格赋值为一个单一的值
                    '''self.e整合_Sheet1[f'{目标列}{报价表整合_Sheet1_maxrow}':f'{目标列}{报价表整合_Sheet1_maxrow+len(col)-int(起始位置)}']\
                        = ws[f'{col_letter}{int(起始位置)}':f'{col_letter}{len(col)}']'''
                # 保存表格
                self.e整合_wb.save(self.e整合_path)
                # 在完成每个sheet的处理后，将当前sheet的名称插入到"报价表整合"表的'来源'列。
                e整合_colstr_来源 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='来源')
                # openpyxl 中，不能直接为一个范围的单元格赋值为一个单一的值
                # self.e整合_Sheet1[f'{e整合_colstr_来源}{报价表整合_Sheet1_maxrow}':f'{e整合_colstr_来源}{报价表整合_Sheet1_maxrow + col_maxlen - int(起始位置)}'] = 报价表名称
                for row in self.e整合_Sheet1[f'{e整合_colstr_来源}{报价表整合_Sheet1_maxrow + 1}':f'{e整合_colstr_来源}{报价表整合_Sheet1_maxrow + 1 + col_maxlen - int(起始位置)}']:
                    row[0].value = f'{报价表名称}/{sheetname}'
                # *保存图片并插入图片到"报价表整合"表的A列，同时写入图片的本地路径
                # 插入图片和命名方式
                命名方式 = '品牌'
                if 系列列号 != None:
                    命名方式 = f'{命名方式}_系列'
                if 名称列号 != None:
                    命名方式 = f'{命名方式}_名称'
                if 货号列号 != None:
                    命名方式 = f'{命名方式}_货号'
                images = ws._images
                # 遍历图像并打印位置信息
                output_folder = os.path.join(self.image_root, 类别, 品牌)
                output_folder_compress = os.path.join(self.image_root, 类别, 品牌, 'thumb')
                if not os.path.exists(output_folder):
                    os.makedirs(output_folder)
                if not os.path.exists(output_folder_compress):
                    os.makedirs(output_folder_compress)
                for index, image in enumerate(images):
                    # 获取图像的左上角行号，即图片所在行
                    img_row = image.anchor._from.row + 1
                    img_column = image.anchor._from.col + 1
                    #img_row = image.anchor.to.row + 1
                    #img_column = image.anchor.to.col + 1
                    #img_column_letter = get_column_letter(img_column)
                    if img_row == 1:
                        # 使用 with open() 是当需要打开一个文件并读写数据时的常见做法
                        # 但在这种情形下，由于 PIL 库中的 save 方法会处理文件的打开和关闭，因此并不需要手动打开文件
                        #with open(os.path.join(output_folder, 'test.png'), 'wb') as img_file:
                        img_pil = Image.open(image.ref).convert("RGB")
                        img_pil.save(os.path.join(output_folder, 'text.png'))
                        continue
                    # TODO: 插入图片前先判断品牌、系列、名称、货号能匹配到几行，如果只有一行，则插入图片。如果不做该功能需要自己在表格中删除重复数据
                    # 获取对应的图片名称
                    图片命名 = 品牌
                    if 系列列号 != None:
                        图片命名 = f'{图片命名}_{self.format_string(ws.cell(row=img_row, column=系列列号).value)}'
                    if 名称列号 != None:
                        图片命名 = f'{图片命名}_{self.format_string(ws.cell(row=img_row, column=名称列号).value)}'
                    if 货号列号 != None:
                        图片命名 = f'{图片命名}_{self.format_string(ws.cell(row=img_row, column=货号列号).value)}'
                    # 图片命名去除换行符
                    # 图片命名 = 图片命名.replace('\n', '').replace('\t', '')
                    图片命名 = re.sub(r'[\n\t]', '', 图片命名)
                    图片命名_thumb = f'{图片命名}_thumb'
                    # 保存图片到本地并按D列的图片名命名
                    img_path = os.path.join(output_folder, f'{图片命名}.png')
                    img_path_compress = os.path.join(output_folder_compress, f'{图片命名_thumb}.png')
                    img_pil = Image.open(image.ref).convert("RGB")
                    if os.path.isfile(img_path) == False:
                        img_pil.save(img_path)
                    # Pillow库中的Image类的save方法可以接受一个'quality'参数来控制保存的图像的质量,img_pil.save(img_path_compress, quality = quality) 实际运行无法实现
                    if os.path.isfile(img_path_compress) == False:
                        new_width, new_height = img_pil.size
                        while True: 
                            # 设置你希望的图片大小，新的宽、高
                            new_width = int(new_width * 0.5)
                            new_height = int(new_height * 0.5)
                            # 使用resize 函数来调整图片尺寸
                            img_pil_resized = img_pil.resize((new_width, new_height))
                            # 保存压缩后的图片
                            img_pil_resized.save(img_path_compress)
                            # 检查新图片的大小
                            if os.path.getsize(img_path_compress) <= 20480: # 图像小于或等于100KB
                                break
                    # 插入图片到表B的A列（行号对应表A同样位置）
                    img = Img(img_path_compress)
                    img.width = 63 # col_ch * 8  col_ch = 8
                    img.height = 61 # row_pt * (4 / 3) row_pt = 46
                    # 修改行高
                    e整合写入行号 = 报价表整合_Sheet1_maxrow + 1 + img_row -int(起始位置)
                    self.e整合_Sheet1.row_dimensions[e整合写入行号].height = 46
                    print(f'写入图片: A 列\t {e整合写入行号} 行')
                    self.e整合_Sheet1.add_image(img, f"{self.e整合_colstr_图片}{e整合写入行号}")
                    self.e整合_Sheet1.cell(row=e整合写入行号, column=self.e整合_colnum_命名方式, value=命名方式)
                    self.e整合_Sheet1.cell(row=e整合写入行号, column=self.e整合_colnum_图片路径, value=img_path)
                    self.e整合_Sheet1.cell(row=e整合写入行号, column=self.e整合_colnum_图片名称, value=图片命名)
                    self.e整合_Sheet1.cell(row=e整合写入行号, column=self.e整合_colnum_品牌, value=品牌)
                # 循环Sheet结束
                self.e整合_wb.save(self.e整合_path)
            # 循环表结束
            self.e整合_wb.save(self.e整合_path)
            # *将当前时间写入"报价表记录"的"记录时间"字段，然后保存整个"报价表记录"
            self.e记录_Sheet1[f'{self.e记录_colstr_记录时间}{e记录行号}'] = datetime.now().strftime("%Y/%m/%d %H:%M")
            self.e记录_wb.save(self.e记录_path)
            
    # TODO: 写入报价表整合时，给每一行两个特征值(两列)(递增)，再加一列类型说明，如果产品是一个系列的，特征值A下拉复制，如果是一个套装的，特征值B下拉复制
    # TODO: 录入报价表的记录时间是一锤子买卖，如果数据出错需要到对照表内进行对照修改或者是代码逻辑的修改
    # TODO: 如果录入的数据是重复的，需要人工做选择，如何选择重复项需要设置
    # TODO: 辨别url域名和指向产品的特征值
    
    def re_operators(self, s):
        # 使用正则表达式找出所有的数字和非数字字符
        numbers = re.findall(r'\d+\.?\d*', s)  # 这将找出所有的数字（包括带小数的）
        numbers = [float(number) for number in numbers]
        numbers.sort(reverse=True)
        operators = re.findall('[^\d.]+', s)  # 这将找出所有的非数字字符
        # 通过将列表转换为集合，再转换回列表，可去除列表中的重复项
        operators_unique = list(set(operators))
        return numbers, operators_unique[0]
    
    # TODO: 格式化数据
    def format_data(self):
        print('开始处理：报价表整合-格式化数据')
        报价表整合_Sheet1_maxrow = self.e整合_Sheet1.max_row
        报价表整合_Sheet1_maxcol = self.e整合_Sheet1.max_column
        
        colstr_产品规格 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='产品规格')
        colstr_产品_长 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='产品-长')
        colstr_产品_宽 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='产品-宽')
        colstr_产品_高 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='产品-高')
        colnum_产品规格 = column_index_from_string(colstr_产品规格) - 1
        colnum_产品_长 = column_index_from_string(colstr_产品_长) - 1
        colnum_产品_宽 = column_index_from_string(colstr_产品_宽) - 1
        colnum_产品_高 = column_index_from_string(colstr_产品_高) - 1
        
        colstr_彩盒规格 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='彩盒规格')
        colstr_彩盒_长 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='彩盒-长')
        colstr_彩盒_宽 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='彩盒-宽')
        colstr_彩盒_高 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='彩盒-高')
        colnum_彩盒规格 = column_index_from_string(colstr_彩盒规格) - 1
        colnum_彩盒_长 = column_index_from_string(colstr_彩盒_长) - 1
        colnum_彩盒_宽 = column_index_from_string(colstr_彩盒_宽) - 1
        colnum_彩盒_高 = column_index_from_string(colstr_彩盒_高) - 1
        
        colstr_外箱规格 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='外箱规格')
        colstr_外箱_长 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='外箱-长')
        colstr_外箱_宽 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='外箱-宽')
        colstr_外箱_高 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='外箱-高')
        colnum_外箱规格 = column_index_from_string(colstr_外箱规格) - 1
        colnum_外箱_长 = column_index_from_string(colstr_外箱_长) - 1
        colnum_外箱_宽 = column_index_from_string(colstr_外箱_宽) - 1
        colnum_外箱_高 = column_index_from_string(colstr_外箱_高) - 1
        
        colstr_毛重净重 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='毛重净重')
        colstr_整箱毛重 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='整箱毛重')
        colstr_整箱净重 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='整箱净重')
        colnum_毛重净重 = column_index_from_string(colstr_毛重净重) - 1
        colnum_整箱毛重 = column_index_from_string(colstr_整箱毛重) - 1
        colnum_整箱净重 = column_index_from_string(colstr_整箱净重) - 1
        
        colstr_名称 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='名称')
        colstr_颗粒数 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='颗粒数')
        colnum_名称 = column_index_from_string(colstr_名称) - 1
        colnum_颗粒数 = column_index_from_string(colstr_颗粒数) - 1
        
        colstr_整箱体积 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='整箱体积')
        colnum_整箱体积 = column_index_from_string(colstr_整箱体积) - 1
        
        colstr_品牌 = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='品牌')
        colnum_品牌 = column_index_from_string(colstr_品牌) - 1
        colstr_编号A = self.find_colname_letter(sheet=self.e整合_Sheet1, rowindex=1, colname='编号A')
        colnum_编号A = column_index_from_string(colstr_编号A) - 1
        A_index_range = self.e整合_Sheet1[f'{A_index_range}{2}:{A_index_range}{报价表整合_Sheet1_maxcol}']
        max_id_dic = {}
        # 计算编号A各个前缀的最大编号
        for cell in row:
            prefix, id = cell.value.split('_')
            id = int(id)
            if prefix not in max_id_dic:
                max_id_dic[prefix] = id
            else:
                if id > max_id_dic[prefix]:
                    max_id_dic[prefix] = id
        
        for i, row in enumerate(self.e整合_Sheet1.iter_rows(min_row=2), start=1):
            if row[colnum_产品规格].value is not None and row[colnum_产品_长].value is None and row[colnum_产品_宽].value is None and row[colnum_产品_高].value is None:
                num, oper = self.re_operators(row[colnum_产品规格].value)
                num_len = len(num)
                if num_len > 0:
                    row[colnum_产品_长].value = num[0] if num_len >= 1 else None
                    row[colnum_产品_宽].value = num[1] if num_len >= 2 else None
                    row[colnum_产品_高].value = num[2] if num_len >= 3 else None
            if row[colnum_彩盒规格].value is not None and row[colnum_彩盒_长].value is None and row[colnum_彩盒_宽].value is None and row[colnum_彩盒_高].value is None:
                num, oper = self.re_operators(row[colnum_彩盒规格].value)
                num_len = len(num)
                if num_len > 0:
                    row[colnum_彩盒_长].value = num[0] if num_len >= 1 else None
                    row[colnum_彩盒_宽].value = num[1] if num_len >= 2 else None
                    row[colnum_彩盒_高].value = num[2] if num_len >= 3 else None
            if row[colnum_外箱规格].value is not None and row[colnum_外箱_长].value is None and row[colnum_外箱_宽].value is None and row[colnum_外箱_高].value is None:
                num, oper = self.re_operators(row[colnum_外箱规格].value)
                num_len = len(num)
                if num_len > 0:
                    row[colnum_外箱_长].value = num[0] if num_len >= 1 else None
                    row[colnum_外箱_宽].value = num[1] if num_len >= 2 else None
                    row[colnum_外箱_高].value = num[2] if num_len >= 3 else None
            if row[colnum_毛重净重].value is not None and row[colnum_整箱毛重].value is None and row[colnum_整箱净重].value is None:
                num, oper = self.re_operators(row[colnum_毛重净重].value)
                num_len = len(num)
                if num_len > 0:
                    row[colnum_整箱毛重].value = num[0] if num_len >= 1 else None
                    row[colnum_整箱净重].value = num[1] if num_len >= 2 else None
                    
            if row[colnum_名称].value is not None and row[colnum_颗粒数].value is None:
                pcs = re.search(r'(\d+)PCS', row[colnum_名称].value)
                pcs_num = pcs.group(1) if pcs else None
                row[colnum_颗粒数].value = pcs_num
            
            # 编号A自动编号
            if row[colnum_编号A].value is None:
                品牌 = row[colnum_品牌].value
                if 品牌 not in max_id_dic:
                    max_id_dic[品牌] = 1
                row[colnum_编号A].value = f'{品牌}_{max_id_dic[品牌]}'
                max_id_dic[品牌] += 1
            
            row[colnum_颗粒数].alignment = Alignment(horizontal='right', vertical='center')
            row[colnum_毛重净重].alignment = Alignment(horizontal='right', vertical='center')
            row[colnum_整箱毛重].alignment = Alignment(horizontal='right', vertical='center')
            row[colnum_整箱净重].alignment = Alignment(horizontal='right', vertical='center')
            row[colnum_整箱体积].alignment = Alignment(horizontal='right', vertical='center')
            
            row[colnum_整箱体积].number_format = '0.000'
            
        self.e整合_wb.save(self.e整合_path)

# 测试代码
if __name__ == '__main__':
    #source_file = r'D:\Code\报价表整合\报价表\明迪积木现货表2024.2.20.xlsx'
    #target_file = r'D:\Code\报价表整合\报价表整合.xlsx'
    
    #ex = ExcelConver(source_file, target_file)
    #ex.stand_execel_contrast()
    #ex.read_excel_contrast()
    #value = ex.read_excel_by_colname_findvalue(sheet_name='展示盒', col_rowindex=1, colname='货号', match_mode='精准匹配', value_rowindex=2)
    #print(value)
    
    '''ex = ExcelConver()
    file_list = ex.list_files_by_type('D:\Code\# 报价表整合\报价表', '.xlsx')
    print(file_list)
    for file_path in file_list:
        base_name = os.path.basename(file_path)
        dir_name = os.path.basename(os.path.dirname(file_path))
        print(f'File name, including extension: {base_name}')
        print(f'Name of the parent directory: {dir_name}')'''
    
    ex = ExcelConver()
    ex.stand_excel()
    ex.recode_excel()
    ex.contrast_data()
    ex.format_data()