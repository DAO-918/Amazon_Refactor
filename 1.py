from openpyxl import load_workbook
import re

# 加载workbook
wb = load_workbook(r'D:\工作-不同步\#5 新品数据\2024.02\开益玩具报价表12月更新.xlsx')


# 打开 Excel 文件
# 获取活动工作表
ws = wb[wb.sheetnames[0]]
pattern = re.compile(r'\d+')
for image in ws._images:
    # 输出图片的位置信息
    print(image.path)
    print(image._path)
    print(image.width)
    print(image.height)
    print(image.anchor._from)
    print(pattern.findall(str(image.anchor._from)))#通过正则获取位置
    xy=pattern.findall(str(image.anchor._from))
    x=xy[0]
    y=xy[2]
